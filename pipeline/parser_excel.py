"""
Parser de ficheiros Excel MQT
Lê ficheiros Excel JSJ formato MQT e extrai artigos
"""
import re
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional
from openpyxl import load_workbook


def _detect_layout(ws) -> dict:
    """
    Detecta layout de colunas lendo linha 14.
    Identifica pares QUANT.+PARCIAL consecutivos para detectar zonas.
    NÃO lê labels de zona do Excel (labels vêm de projects.zona_config).
    Fallback: se não encontra pares, assume 1 zona (QUANT. simples).
    
    Retorna dict com índices 0-based:
    {
        quant_cols: [6, 8, 10],  # índices das colunas QUANT. que precedem PARCIAL
        quant_total: 13,
        preco_unit: 14,
        total_eur: 15,
        num_zonas: 3
    }
    """
    header_row = ws[14]
    quant_cols = []
    quant_total_col = None
    preco_unit_col = None
    total_eur_col = None
    prev_was_quant = False
    prev_quant_idx = None

    for i, cell in enumerate(header_row[:24]):
        val = str(cell.value or '').strip().upper().replace('\n', ' ')
        if val in ('QUANT.', 'QUANT'):
            prev_was_quant = True
            prev_quant_idx = i
        elif val == 'PARCIAL':
            if prev_was_quant and prev_quant_idx is not None:
                quant_cols.append(prev_quant_idx)
            prev_was_quant = False
            prev_quant_idx = None
        elif 'QUANT. TOTAL' in val or 'QUANT.TOTAL' in val or 'QUANT TOTAL' in val:
            quant_total_col = i
            prev_was_quant = False
        elif 'P. UNIT' in val or 'P.UNIT' in val:
            preco_unit_col = i
            prev_was_quant = False
        elif val in ('TOTAL €', 'TOTAL'):
            total_eur_col = i
            prev_was_quant = False
        else:
            prev_was_quant = False
            prev_quant_idx = None

    if not quant_cols:
        for i, cell in enumerate(header_row[:24]):
            val = str(cell.value or '').strip().upper()
            if val in ('QUANT.', 'QUANT'):
                quant_cols = [i]
                break

    return {
        'quant_cols': quant_cols,
        'quant_total': quant_total_col,
        'preco_unit': preco_unit_col,
        'total_eur': total_eur_col,
        'num_zonas': len(quant_cols) if quant_cols else 1,
    }


def parse_mqt(excel_path) -> Dict:
    """
    Faz parse de um ficheiro Excel MQT JSJ
    
    Especificações:
    - Sheet: '02.MQT'
    - Header: linha 14
    - Dados: linhas 15 a 153
    - Artigos reais: col D é string formato 'X.Y.Z' (não float)
    
    Args:
        excel_path: Caminho para ficheiro Excel (str) ou file-like object (upload)
        
    Returns:
        Dict com:
        {
            'artigos': [...],  # lista de dicts com artigos
            'num_zonas': 3,    # número de zonas detectadas no layout
        }
    """
    if hasattr(excel_path, 'read'):
        # file-like object (Streamlit UploadedFile)
        excel_path.seek(0)
        wb = load_workbook(BytesIO(excel_path.read()), data_only=True)
    else:
        # path string
        wb = load_workbook(excel_path, data_only=True)
    
    # Verificar se sheet existe
    sheet_name = '02.MQT'
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' não encontrada. Sheets disponíveis: {wb.sheetnames}")
    
    ws = wb[sheet_name]
    
    # Detectar layout automaticamente
    layout = _detect_layout(ws)
    num_zonas = layout['num_zonas']

    rows_out = []
    current_subcap = None
    current_classe = None
    last_artigo_idx = None  # índice do último artigo em rows_out

    for row in ws.iter_rows(min_row=15, max_row=ws.max_row, values_only=True):
        cod_raw = row[3]
        desc_raw = row[4]
        unidade_raw = row[5]
        desc = str(desc_raw).strip() if desc_raw else None
        unidade = str(unidade_raw).strip() if unidade_raw else None

        # Linha vazia — ignorar
        if cod_raw is None and not desc:
            continue

        # Linha SPEC — juntar ao artigo anterior
        if cod_raw is None and desc:
            if last_artigo_idx is not None:
                rows_out[last_artigo_idx]['especificacao'] = desc
            continue

        cod_str = str(cod_raw).strip().strip("'\"")

        # Normalizar floats (3.0 → "3", 5.1 → "5.1")
        try:
            f = float(cod_str)
            parts_normalized = str(f).rstrip('0').rstrip('.')
            cod_str = parts_normalized
        except Exception:
            pass

        parts = cod_str.split('.')
        num_parts = len(parts)

        # Ler quantidades usando valores raw
        quant_a = _val_to_float(row[layout['quant_cols'][0]]) if len(layout['quant_cols']) > 0 else None
        quant_b = _val_to_float(row[layout['quant_cols'][1]]) if len(layout['quant_cols']) > 1 else None
        quant_c = _val_to_float(row[layout['quant_cols'][2]]) if len(layout['quant_cols']) > 2 else None
        quant_d = _val_to_float(row[layout['quant_cols'][3]]) if len(layout['quant_cols']) > 3 else None
        col_qt = layout.get('quant_total')
        quant_total = _val_to_float(row[col_qt]) if col_qt is not None else None
        if quant_total is None:
            quant_total = round(sum(q or 0 for q in [quant_a, quant_b, quant_c, quant_d]), 4) or None

        # CAP — 1 segmento
        if num_parts == 1:
            current_subcap = None
            current_classe = None
            artigo = {
                'nivel': 1,
                'artigo_cod': cod_str,
                'descricao': desc,
                'unidade': None,
                'especificacao': None,
                'classe_material': None,
                'elemento_sufixo': None,
                'is_nivel4': False,
                'capitulo': cod_str,
                'subcapitulo': None,
                'sufixo': None,
                'quant_a': None, 'quant_b': None,
                'quant_c': None, 'quant_d': None, 'quant_total': None,
            }
            rows_out.append(artigo)
            last_artigo_idx = len(rows_out) - 1
            continue

        # SUBCAP — 2 segmentos
        if num_parts == 2:
            current_subcap = cod_str
            current_classe = _extrair_classe(desc) if desc else None

            if unidade:
                # Subcapítulo com unidade → tratar como artigo nivel=2
                artigo = {
                    'nivel': 2,
                    'artigo_cod': cod_str,
                    'descricao': desc,
                    'unidade': unidade,
                    'especificacao': None,
                    'classe_material': current_classe,
                    'elemento_sufixo': parts[1] if len(parts) > 1 else None,
                    'is_nivel4': False,
                    'capitulo': parts[0],
                    'subcapitulo': cod_str,
                    'sufixo': parts[1] if len(parts) > 1 else None,
                    'quant_a': quant_a, 'quant_b': quant_b,
                    'quant_c': quant_c, 'quant_d': quant_d,
                    'quant_total': quant_total,
                }
            else:
                # Subcapítulo título — sem quantidades
                artigo = {
                    'nivel': 2,
                    'artigo_cod': cod_str,
                    'descricao': desc,
                    'unidade': None,
                    'especificacao': None,
                    'classe_material': current_classe,
                    'elemento_sufixo': None,
                    'is_nivel4': False,
                    'capitulo': parts[0],
                    'subcapitulo': cod_str,
                    'sufixo': None,
                    'quant_a': None, 'quant_b': None,
                    'quant_c': None, 'quant_d': None, 'quant_total': None,
                }
            rows_out.append(artigo)
            last_artigo_idx = len(rows_out) - 1
            continue

        # ARTIGO — 3+ segmentos
        is_nivel4 = num_parts >= 4
        nivel = 4 if is_nivel4 else 3
        elemento_sufixo = parts[2] if len(parts) >= 3 else None
        sufixo = '.'.join(parts[2:]) if len(parts) >= 3 else None

        artigo = {
            'nivel': nivel,
            'artigo_cod': cod_str,
            'descricao': desc,
            'unidade': unidade,
            'especificacao': None,
            'classe_material': current_classe,
            'elemento_sufixo': elemento_sufixo,
            'is_nivel4': is_nivel4,
            'capitulo': parts[0],
            'subcapitulo': current_subcap,
            'sufixo': sufixo,
            'quant_a': quant_a, 'quant_b': quant_b,
            'quant_c': quant_c, 'quant_d': quant_d,
            'quant_total': quant_total,
        }
        rows_out.append(artigo)
        last_artigo_idx = len(rows_out) - 1

    wb.close()

    n1 = sum(1 for r in rows_out if r['nivel'] == 1)
    n2 = sum(1 for r in rows_out if r['nivel'] == 2)
    n3 = sum(1 for r in rows_out if r['nivel'] >= 3)
    print(f"   ✓ {len(rows_out)} linhas parseadas (cap={n1}, subcap={n2}, artigos={n3})")
    print(f"   ℹ {num_zonas} zona(s) detectada(s) no layout")

    return {
        'artigos': rows_out,
        'num_zonas': num_zonas,
    }


def extract_capitulo_info(artigo_cod: str) -> Dict[str, Optional[str]]:
    """
    Extrai informação do código de artigo
    
    Regras:
    - capitulo   = 1º segmento         ex: '5.5.1'   → '5'
    - subcapitulo = 1º+2º segmento     ex: '5.5.1'   → '5.5'
    - sufixo     = restantes segmentos ex: '5.5.1'   → '1'
                                        ex: '6.2.4.1' → '4.1'
                                        ex: '15.3.1'  → '3.1'
    - elemento_sufixo = 3º segmento    ex: '5.5.1'   → '1'
                                        ex: '6.2.4.1' → '4'
                                        ex: '15.3.1'  → '1'
    
    Args:
        artigo_cod: código do artigo (ex: "5.5.4", "6.2.4.1", "15.3.1")
        
    Returns:
        Dict com capitulo, subcapitulo, sufixo, elemento_sufixo (strings ou None se inválido)
    """
    partes = artigo_cod.split('.')
    
    # Mínimo 3 segmentos para ser artigo válido
    if len(partes) < 3:
        return {
            'capitulo': None,
            'subcapitulo': None,
            'sufixo': None,
            'elemento_sufixo': None
        }
    
    try:
        capitulo = partes[0]  # '5' ou '15'
        subcapitulo = f"{partes[0]}.{partes[1]}"  # '5.5' ou '15.3'
        sufixo = '.'.join(partes[2:])  # '1' ou '4.1' ou '3.1'
        elemento_sufixo = partes[2]  # 3º segmento apenas: '1' ou '4'
        
        return {
            'capitulo': capitulo,
            'subcapitulo': subcapitulo,
            'sufixo': sufixo,
            'elemento_sufixo': elemento_sufixo
        }
    except Exception:
        return {
            'capitulo': None,
            'subcapitulo': None,
            'sufixo': None,
            'elemento_sufixo': None
        }


def extract_classe_material(descricao: str) -> Optional[str]:
    """
    Extrai classe de material da descrição (ex: C30/37, C20/25)
    
    Args:
        descricao: texto da descrição do artigo
        
    Returns:
        String com classe (ex: 'C30/37') ou None se não encontrado
    """
    if not descricao:
        return None
    
    # Regex para capturar padrão CXX/XX
    pattern = r'C\d+/\d+'
    match = re.search(pattern, descricao)
    
    if match:
        return match.group(0)
    
    return None


def _get_cell_string(cell) -> str:
    """
    Obtém valor de célula como string (stripped)
    
    Args:
        cell: objeto Cell do openpyxl
        
    Returns:
        String (vazia se None)
    """
    value = cell.value
    if value is None:
        return ""
    return str(value).strip()


def _get_cell_float(cell) -> Optional[float]:
    """
    Obtém valor de célula como float

    Args:
        cell: objeto Cell do openpyxl

    Returns:
        Float ou None se vazio/inválido
    """
    value = cell.value
    if value is None or value == "":
        return None

    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _val_to_float(v) -> Optional[float]:
    """
    Converte valor raw (de iter_rows values_only=True) para float.

    Args:
        v: valor raw da célula

    Returns:
        Float ou None se vazio/inválido
    """
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _extrair_classe(desc: str) -> Optional[str]:
    """
    Extrai classe de material do nome do subcapítulo.
    Ex: 'Betão C30/37'   → 'C30/37'
        'Aço A500NR SD'  → 'A500NR SD'
        'Classe A2 (...)' → 'A2'

    Args:
        desc: texto da descrição

    Returns:
        String com classe ou None
    """
    if not desc:
        return None
    # Betão: C12/15, C20/25, C30/37, C35/45...
    m = re.search(r'C\d+/\d+', desc)
    if m:
        return m.group()
    # Aço: A500NR SD, A400NR, A500NR...
    m = re.search(r'A\d+[A-Z\s]*(?:SD)?', desc)
    if m:
        return m.group().strip()
    # Cofragem: Classe A1, A2, A3, A4
    m = re.search(r'Classe (A\d)', desc)
    if m:
        return m.group(1)
    return None


def validate_artigos(artigos: List[Dict]) -> bool:
    """
    Valida lista de artigos parseados
    
    Args:
        artigos: lista de dicts com artigos
        
    Returns:
        True se válido, False com warnings caso contrário
    """
    if not artigos:
        print("⚠️  Nenhum artigo encontrado")
        return False
    
    issues = 0
    
    for i, artigo in enumerate(artigos):
        # Validar campos obrigatórios
        if not artigo.get('artigo_cod'):
            print(f"⚠️  Linha {i+1}: artigo_cod vazio")
            issues += 1
        
        if not artigo.get('descricao'):
            print(f"⚠️  Linha {i+1}: descricao vazia para {artigo.get('artigo_cod')}")
            issues += 1
        
        # Validar quant_total = quant_a + quant_b + quant_c (se todos definidos)
        qa = artigo.get('quant_a') or 0
        qb = artigo.get('quant_b') or 0
        qc = artigo.get('quant_c') or 0
        qt = artigo.get('quant_total') or 0
        
        soma = qa + qb + qc
        if abs(qt - soma) > 0.01 and qt != 0:
            print(f"⚠️  {artigo['artigo_cod']}: quant_total ({qt}) ≠ soma ({soma})")
            issues += 1
    
    if issues > 0:
        print(f"⚠️  {issues} problemas encontrados na validação")
        return False
    
    print(f"✅ Validação OK: {len(artigos)} artigos válidos")
    return True
